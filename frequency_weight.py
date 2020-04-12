from konlpy.tag import Kkma
from openpyxl import load_workbook


def importCSV(infile):
#파일 열고 읽기
    with open(infile, 'r') as input_file:
        read_file = input_file.readlines()

    return filename, read_file

def tokenization(csvfile):
#csv 다듬기(어절 단위)
    total = []
    for line in csvfile:
        a = line.split()
        for word in a:
            word = word.strip()
            total.append(word)
    total = total[1:]
    return total

def posTagging(tokenized):
#형태소 분석기 돌리기
    kkma = Kkma()

    total_tag = []
    for item in tokenized:
        a = kkma.pos(item)
        for aa in a:
            total_tag.append(aa)
    return total_tag

def stopWords(postagged):
#필요없는 품사의 단어들 제외하기
    stopwords = ['VXV','VXA','VCP','VCN','MDN','MDT','NR','MAC','JKS','JKC','JKG','JKO','JKM','JKI','JKQ','JC','JX','EPH','EPT','EPP','EFN','EFQ','EFO','EFA','EFI','EFR','ECE','ECS','ECD','ETN','ETD','SF','SE','SS','SP','SO','SW','OH','ON','UN']
    total_tag_cleansed = []
    for tag in postagged:
        if tag[1] not in stopwords:
            total_tag_cleansed.append(tag[0])
    return total_tag_cleansed


def countingFrequency(cleansed):
#단어 빈도수 세기
    freq = {}
    for words in cleansed:
        freq[words] = freq.get(words, 0) + 1
    return freq

def dict2list(dictionary):
#빈도수 사전 리스트로 전환
    result = []
    for key in dictionary:
        a = []
        a.append(key)
        a.append(dictionary[key])
        result.append(a)
    return result

def Sort(result):
#리스트 대락적으로 분류
    result_sorted = []
    for r in result:
        if int(r[1]) > 20:

            if int(r[1]) > 50:
                result_sorted.insert(0, r)
            else:
                result_sorted.insert(1, r)


        else:
            if int(r[1]) < 2:
                pass
            else:
                result_sorted.insert(-1, r)
    return result_sorted


def exportXLSX(filename, result_sorted):
#빈도 저장할 엑셀 파일 만들기
    wb = load_workbook("빈도정리_가중치.xlsx")
    ws = wb.create_sheet(title= filename)

    tot = 1
    weight = float(input("Weight: ")) #가중치 입력 받기
    for rr in result_sorted:
        ws.cell(row=tot, column=1).value = rr[0]
        ws.cell(row=tot, column=2).value = int(rr[1])*weight #가중치만큼 빈도 곱하기
        tot += 1

    wb.save("빈도정리_가중치.xlsx")


def main(filename, infile):
    csvfile = importCSV(infile)
    tokenized = tokenization(csvfile)
    postagged = posTagging(tokenized)
    cleansed = stopWords(postagged)
    freq = countingFrequency(cleansed)
    total_list = dict2list(freq)
    sorted_list = Sort(total_list)
    exportXLSX(filename, sorted_list)









#input으로 파일 이름 받기
filename = input('filename: ')
infile = 'wgang_'+ filename + '.csv'
main(filename, infile)